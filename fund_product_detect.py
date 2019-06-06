from io import StringIO
from io import open
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, process_pdf
import pickle
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import utils
import time
import os
from openpyxl import workbook
from openpyxl import load_workbook
from pymongo import MongoClient

conn = MongoClient('localhost', 27017)
db = conn.mydb
my_set = db.fund_history

def read_pdf(pdf_path):
    with open(pdf_path, "rb") as pdf:
        # resource manager
        rsrcmgr = PDFResourceManager()
        retstr = StringIO()
        laparams = LAParams()
        # device
        device = TextConverter(rsrcmgr, retstr, laparams=laparams)
        process_pdf(rsrcmgr, device, pdf)
        device.close()
        content = retstr.getvalue()
        retstr.close()

        #fencoding = chardet.detect(content)
        #print(fencoding)
        # print(type(content), flush=True)
        # lines = str(content).split("\n")
        # return lines
        return str(content)

def re_find(key_word, content):
    try:
        str_list = list(key_word)
        temp = u'\s*'.join(str_list)
        findword = temp
        #print(findword)
        # pattern = re.compile(findword)
        # find_results = pattern.findall(content)

        # return find_results

        return re.findall(findword, content)
    except Exception as e:
        return []


def compare_pdf(contract_pdf_path, approval_pdf_path):
    company, product = None, None
    contract_content = read_pdf(contract_pdf_path)
    
    #findword = u'基金或本基金：(.*)'
    findword = u'基金或本基金：指([\s\S]*?)2'
    pattern = re.compile(findword)
    results = pattern.findall(contract_content)
    print(results)

    if len(results) != 0:
        findword = results[0].strip()
        approval_content = read_pdf(approval_pdf_path)
        #print(approval_content)
        str_list = list(findword)
        temp = u'\s*'.join(str_list)
        findword = temp
        #print(findword)
        results = re.findall(findword, approval_content)
        print(results)
        if len(results) != 0:
            product = results[0].strip()

    findword = u'基金管理人：指(.*)'
    pattern = re.compile(findword)
    results = pattern.findall(contract_content)
    if len(results) != 0:
        company = results[0].strip()
        print(company)

    return company, product

# company, product = compare_pdf('D:/IPA/external/基金对比资料/基金对比资料/广发9只基金/广发中证军工ETF联接C/基金合同.pdf', \
#     'D:/IPA/external/基金对比资料/基金对比资料/广发9只基金/广发中证军工ETF联接C/招募说明书201801.pdf')

def create_excel(company, product, code, name, department):  
    template_path = utils.get_working_dir() + '/基金审核.xlsx'
    w_workbook = load_workbook(template_path)
    sheets = w_workbook.sheetnames
    w_sheet = w_workbook[sheets[0]]

    w_sheet.cell(row = 6, column = 2).value = product.replace(" ", "").replace("\n", "")

    title = '微众银行合规审查申请表  \n（合规评审意见）\n                                编号：' + code
    w_sheet.cell(row = 2, column = 1).value = title

    content = '   合规审查意见（法律合规部填写）：\n \n    根据业务侧提供的相关资料，本产品无重大法律合规风险，合规意见如下：\n   \n    \
    一、' + company + '拥有中国证券监督管理委员会颁发的“经营证券期货业务许可证”，发行方符合《关于规范商业银行代理销售业务的通知》所允许的商业银行代销产品发行机构，我行具有公募基金代销资格。\n    \n     \
    二、产品是经中国证券监督管理委员会备案的公募基金产品，备案通过日期为2013年。\n\n     \
    三、请业务侧严格按照《关于规范商业银行代理销售业务的通知》、《证券投资基金销售管理办法》开展产品销售活动。产品宣传页(如有)需另行送审法律合规部。\n     \n    \
    四、产品上架后，做好产品存续期管理，及时更新产品重要公告。\n   '

    # print(repr(w_sheet.cell(row = 2, column = 1).value))

    w_sheet.cell(row = 7, column = 1).value = content

    date = time.strftime("%Y-%m-%d", time.localtime())
    w_sheet.cell(row = 18, column = 2).value = date
    w_sheet.cell(row = 3, column = 4).value = name
    w_sheet.cell(row = 3, column = 2).value = department

    file_name = ('合规评审表' + code + '(' + product + ')').strip().replace('\n','')
    save_path = utils.get_working_dir() + '/' + file_name + '.xlsx'
    temp = save_path.strip().replace('\n','')
    w_workbook.save(temp)

    return temp, file_name


def run(contract_pdf_path, approval_pdf_path, code, name, department):
    company, product = compare_pdf(contract_pdf_path, approval_pdf_path)
    result_path, file_name = create_excel(company, product, code, name, department) 

    date = time.strftime("%Y-%m-%d", time.localtime())
    my_set.insert({'name': file_name, \
    'result_path': result_path, \
    'contract_pdf_path': contract_pdf_path, \
    'approval_pdf_path': approval_pdf_path, \
    'date': date})

    return result_path, file_name

def search_history(text, date_start, date_end):
    for item in my_set.find( {'$and': [ {'date': {'$gte': date_start, '$lt': date_end}}, \
    {'name':{'$regex': text}} \
    ] } ):

        print(item['name'])

if __name__ == '__main__':
    result_path, file_name = run('D:/IPA/external/基金对比资料/基金对比资料/广发9只基金/广发中证军工ETF联接C/基金合同.pdf', \
        'D:/IPA/external/基金对比资料/基金对比资料/广发9只基金/广发中证军工ETF联接C/招募说明书201801.pdf', \
        '11223344', '123', '456')

    print(result_path)
    print(file_name)