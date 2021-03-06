from io import StringIO
from io import open
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, process_pdf
import pickle
import re
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import utils
import time
import os
from openpyxl import workbook
from openpyxl import load_workbook
import win32com.client
import textract
from pymongo import MongoClient
 
conn = MongoClient('localhost', 27017)
db = conn.mydb
my_set = db.insurance_history

products_info = {}

def catch_data(driver, url):    
    global products_info
    
    driver.get(url)
    driver.implicitly_wait(30)

    driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
    #table = driver.find_element_by_xpath("//table[@class='orglist_main']")
    elems = driver.find_elements_by_xpath("//td[@class='orglist_td']/a")
    #print(len(elems))

    for elem in elems:
        company = elem.get_attribute('innerText')
        print(company)

        if company in products_info:
            continue
        
        elem.click()
        driver.switch_to_window(driver.window_handles[1])

        # js = "var q=document.documentElement.scrollTop=10000"
        # driver.execute_script(js)
        time.sleep(1)
        
        products = driver.find_elements_by_xpath("//td[@style=' text-align:left;']")
        product_list = []
        #print(len(products))
        for i in products:
            product_list.append(i.get_attribute('innerText'))

        time.sleep(1)
        driver.close()    
        time.sleep(1)
        driver.switch_to_window(driver.window_handles[0])
        time.sleep(1)
        driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))

        if company in products_info:
            products_info[company].extend(product_list)
        else:
            products_info[company] = product_list

        time.sleep(1)
        
    return products_info

def read_file(path):
    filename, file_extension = os.path.splitext(path)
    if file_extension == '.pdf':
        with open(path, "rb") as pdf:
            # resource manager
            rsrcmgr = PDFResourceManager()
            retstr = StringIO()
            laparams = LAParams()
            # device
            device = TextConverter(rsrcmgr, retstr, laparams=laparams)
            process_pdf(rsrcmgr, device, pdf)
            device.close()
            content = retstr.getvalue()
            retstr.close()

            return str(content)
    elif file_extension == '.doc':
        word = win32com.client.Dispatch("Word.Application")
        word.visible = False
        wb = word.Documents.Open(path)
        doc = word.ActiveDocument
        return doc.Range().Text
    elif file_extension == '.docx':
        temp = textract.process(path)
        return temp.decode()
 
def re_find(key_word, content):
    try:
        str_list = list(key_word)
        temp = u'\s*'.join(str_list)
        findword = temp
        #print(findword)
        # pattern = re.compile(findword)
        # find_results = pattern.findall(content)

        # return find_results

        return re.findall(findword, content)
    except Exception as e:
        return []

def get_produt_info(path):
    product_company, product_name = None, None
    content = read_file(path)
    products_data = pickle.load(open("product_info.p", "rb"))

    for company in products_data:
        if company == '1' or company == '3':
            continue

        find_company = re_find(company, content)
        for i in find_company:
            print(i)
            product_company = i
            products_list = products_data[company]
            for product in products_list:
                find_product = re_find(product, content)
                if len(find_product) != 0:
                    print(find_product)
                    product_name = find_product[0]
                    break
            break

    return product_company, product_name

# 数据抓取
# driver = utils.ChromeBrowser()
# result = catch_data(driver, 'http://bxjg.circ.gov.cn/tabid/6757/Default.aspx')
# driver.quit()

# driver = utils.ChromeBrowser()
# result = catch_data(driver, 'http://bxjg.circ.gov.cn/tabid/6758/Default.aspx')
# driver.quit()
# pickle.dump(result, open("product_info.p", "wb"))

def auto_catch_data():
    driver = utils.ChromeBrowser()
    try:
        result = catch_data(driver, 'http://bxjg.circ.gov.cn/tabid/6757/Default.aspx')
        driver.quit()

        driver = utils.ChromeBrowser()
        result = catch_data(driver, 'http://bxjg.circ.gov.cn/tabid/6758/Default.aspx')
        driver.quit()
        pickle.dump(result, open("product_info.p", "wb"))
    except Exception as e:
        driver.quit()
        print('发生异常等待5秒自动重新调用函数.......')
        time.sleep(5)
        auto_catch_data()

def create_excel(company, product, code, name, department):
    template_path = utils.get_working_dir() + '/保险审核.xlsx'
    w_workbook = load_workbook(template_path)
    sheets = w_workbook.sheetnames
    w_sheet = w_workbook[sheets[0]]

    title = '微众银行合规审查申请表  \n（合规评审意见）\n                                编号：' + code
    w_sheet.cell(row = 2, column = 1).value = title

    w_sheet.cell(row = 6, column = 2).value = product.replace(" ", "").replace("\n", "")
    content = '合规审查意见（法律合规部填写）：\n \n   根据业务侧提供的相关资料，本产品无重大法律合规风险，合规意见如下：\n   \n    \
        一、本产品发行人为' + company + '，具有保监会颁发保险公司法人许可证，根据《关于规范商业银行代理销售业务的通知》（银监发2016[24]号）规定：商业银行可接受国务院证券监督管理机构管理并持有金融牌照的金融机构委托，在本行渠道向客户推介、销售合作机构依法发行的金融产品。\n      \n     \
        二、经业务侧尽调，可于银保监会网站查询该产品。属于依法发行的保险产品。\n\n     \
        三、合规提示\n    1.交互页面及关于产品功能的描述需经合作方确认；2.投诉、理赔需由合作方负责，并在合同中约定双方权责义务关系。\n    \n'
    w_sheet.cell(row = 7, column = 1).value = content

    date = time.strftime("%Y-%m-%d", time.localtime()) 
    w_sheet.cell(row = 19, column = 2).value = date
    w_sheet.cell(row = 3, column = 4).value = name
    w_sheet.cell(row = 3, column = 2).value = department

    file_name = ('合规评审表' + code + '(' + product + ')').strip().replace('\n','')
    save_path = utils.get_working_dir() + '/' + file_name + '.xlsx'
    temp = save_path.strip().replace('\n','')
    w_workbook.save(temp)

    return temp, file_name
 
def run(file_path, code, name, department):
    company, product = get_produt_info(file_path)   
    result_path, file_name = create_excel(company, product, code, name, department)

    date = time.strftime("%Y-%m-%d", time.localtime())
    my_set.insert({'name': file_name, \
        'result_path': result_path, \
        'file_path': file_path, \
        'date': date})

    return result_path, file_name

def search_history(text, date_start, date_end):
    for item in my_set.find( {'$and': [ {'date': {'$gte': date_start, '$lt': date_end}}, \
    {'name':{'$regex': text}} \
    ] } ):

        print(item['name'])    

if __name__ == '__main__':
    # company, product = get_produt_info('d:/test2.pdf')   
    # print(create_excel(company, product, '123456', 'asdf', 'qwer'))   
    result_path, file_name = run('d:/test2.pdf', '123456', 'asdf', 'qwer')
    print(result_path, file_name)