import requests
from multiprocessing import Queue
# from multiprocessing import Queue
# from six.moves.queue import Queue
# from idna import idnadata
from lxml import etree
# from urllib.request import urlretrieve
import time
# from bs4 import  BeautifulSoup
from openpyxl import workbook
from openpyxl import load_workbook
import pickle
import os

def run(username, password):
    catch_data = []

    # if os.path.exists('catch_data.p'):
    #     with open("catch_data.p", "rb") as f:
    #         catch_data = pickle.load(f)

    headers = {
    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
    }

    login_data = {
        'commit': 'Sign in',
        'utf8': '%E2%9C%93',
        'login': username,
        'password': password
    }

    with requests.Session() as session:
        url = "https://github.com/session"
        session_result = session.get(url, headers=headers)
        session_html = etree.HTML(session_result.content)
        temp = session_html.xpath('//input[@name="authenticity_token"]')
        # print(len(temp))
        for item in temp:
            # print(item.text, flush=True)
            # print(etree.tostring(item, pretty_print=True))
            # print(item.get('value'))
            login_data['authenticity_token'] = item.get('value')
            break

        # soup = BeautifulSoup(session_result.content, 'html5lib')
        # login_data['authenticity_token'] = soup.find('input', attrs={'name': 'authenticity_token'})['value']
        # print(login_data['authenticity_token'])

        r = session.post(url, data=login_data, headers=headers)
        # print(r.content)

        url = None
        count = 0
        url = 'https://github.com/WeBankFinTech/FATE/stargazers'
        while True:
            result = session.get(url)
            html = etree.HTML(result.content)

            user_list = html.xpath('//*[@class="follow-list-name"]/span/a')
            
            for user_name in user_list:
                record = None
                # for item in catch_data:
                #     if item["username"] == user_name.text:
                #         record = item

                # if record is not None:
                #     continue

                count += 1
                # print(item.text)
                user_page = 'https://github.com/' + user_name.text
                # print(user_page)
                user_page_result = session.get(user_page)
                user_page_html = etree.HTML(user_page_result.content)

                # catche_history.append(user_name.text)

                email_list = user_page_html.xpath('//*[@class="u-email "]')
                # print(len(email_list))
                item = {}
                item['username'] = user_name.text
                item['email'] = None
                for email in email_list:
                    print(email.text, flush=True)
                    # item['username'] = user_name.text
                    item['email'] = email.text
                catch_data.append(item)

                # time.sleep(1)    

            button_list = html.xpath('//div[@class="paginate-container"]/div/a')
            # print(len(button_list))

            have_next_page = False
            for button in button_list:
                if button.text == 'Next':
                    have_next_page = True
                    url = button.get('href')
                    print(url)

            if not have_next_page:
                with open("catch_data.p", "wb") as f:
                    pickle.dump(catch_data, f)
                return catch_data

def save_excel(data):
    template_path = 'template.xlsx'
    w_workbook = load_workbook(template_path)
    sheets = w_workbook.sheetnames
    w_sheet = w_workbook[sheets[0]]

    write_row = 1
    for item in data:
        if item['email'] is None:
            continue

        write_row += 1
        w_sheet.cell(row = write_row, column = 1).value = write_row - 1
        w_sheet.cell(row = write_row, column = 2).value = item['username']
        w_sheet.cell(row = write_row, column = 3).value = item['email']

    w_workbook.save('result.xlsx')

result = run('cxn8801@gmail.com', 'cxn8801cxn')
save_excel(result)